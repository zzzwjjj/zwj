import pandas as pd
import torch
import torch.nn as nn
from sklearn.preprocessing import StandardScaler
import numpy as np
import os
import openpyxl
# 加载数据
def delete_excel_file(file_path):
    try:
        # 获取当前文件所在的目录路径
        # 检查文件是否存在，如果存在则删除
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        print(f"删除文件时出错: {e}")
def go():

    data_path = './Analyze.xlsx'
    if os.path.exists(data_path):
        data = pd.read_excel(data_path)
        # 1. 移除'name'列
        data = data.drop(columns=['name'])
        # 2. 转换'Package'和'Environment'列
        # 先定义一个函数来处理这两列
        def split_comma_separated_values(row, column_name):
            values = row[column_name].split(',')
            return [int(value.strip()) for value in values]

        # 应用转换
        package_columns = data.apply(lambda row: split_comma_separated_values(row, 'Package'), axis=1)
        environment_columns = data.apply(lambda row: split_comma_separated_values(row, 'Environment'), axis=1)

        # 将这些列扩展到原始DataFrame中
        max_package_len = max(len(values) for values in package_columns)
        max_environment_len = max(len(values) for values in environment_columns)

        for i in range(max_package_len):
            data[f'Package_{i+1}'] = package_columns.apply(lambda x: x[i] if i < len(x) else 0)

        for i in range(max_environment_len):
            data[f'Environment_{i+1}'] = environment_columns.apply(lambda x: x[i] if i < len(x) else 0)

        # 移除原始的'Package'和'Environment'列
        data = data.drop(columns=['Package', 'Environment'])

        # 3. 重命名标签列
        data = data.rename(columns={'Unnamed: 11': 'label'})

        # 4. 特征和标签分离
        X = data.drop(columns=['label'])



        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        X_reshaped = np.reshape(X_scaled, (X_scaled.shape[0], 1, X_scaled.shape[1]))
        inputs = torch.tensor(X_reshaped,dtype=torch.float32)


        class GRUModel(nn.Module):
            def __init__(self, input_size, hidden_size, num_layers, num_classes):
                super(GRUModel, self).__init__()
                self.gru = nn.GRU(input_size, hidden_size, num_layers, batch_first=True)
                self.fc = nn.Linear(hidden_size, num_classes)

            def forward(self, x):
                _, h_n = self.gru(x)  # GRU的输出是(output, h_n)
                h_n = h_n[-1, :, :]  # 取最后一层的h_n
                out = self.fc(h_n)
                return out
        dict_list=['best_model_weights_1.pth','best_model_weights_2.pth','best_model_weights_3.pth','best_model_weights_4.pth','best_model_weights_5.pth','best_model_weights_6.pth','best_model_weights_7.pth','best_model_weights_8.pth']
        b=1
        label_name_dict = {label: [] for label in range(10)}  # 假设标签范围为0到9
        for a in dict_list:
            model = GRUModel(input_size=inputs.shape[2], hidden_size=32, num_layers=1, num_classes=9)
            file_path = f'./utils/{a}'
            model.load_state_dict(torch.load(file_path))
            model.eval()
            with torch.no_grad():  # 禁用梯度计算
                outputs = model(inputs)
                _, predicted = torch.max(outputs, 1)


            # predicted_data_path = './inference.xlsx'
            predicted_data = pd.read_excel(data_path)
            predicted_data['label'] = predicted.numpy()

            predicted_data.to_excel('./predicted_data.xlsx')



            # 将名称添加到相应的键中
            for index, row in predicted_data.iterrows():
                label = row['label']
                # if label ==0:
                #     name = row['name']
                #     label_name_dict[label].append(name)
                # else:
                #     name= row['name']
                #     label_name_dict[b].append(name)
                if label ==1:
                    name= row['name']
                    label_name_dict[b].append(name)
            # print(label_name_dict[b])
            # print(len(label_name_dict[b]))
            b=b+1

        workbook = openpyxl.load_workbook('Analyze.xlsx')
        sheet = workbook.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        for row in range(2, max_row + 1):
            for column in range(1, max_column + 1):
                cell = sheet.cell(row=row, column=column)
                cell.value = None
        workbook.save('Analyze.xlsx')
    else:
        label_name_dict = []
    return label_name_dict
